from __future__ import annotations

from dataclasses import dataclass
from typing import List
import io
import logging
import os
import re

from fastapi import HTTPException, UploadFile

from services.document_text import extract_text_from_document
from services.extraction.normalize import normalize_text

logger = logging.getLogger(__name__)

# optional dependencies (openpyxl for xlsx)
try:
    import openpyxl
except Exception:
    openpyxl = None


MAX_FILES = 25
MAX_FILE_BYTES = 15 * 1024 * 1024  # 15MB на файл (можно поменять)
ALLOWED_EXT = {".pdf", ".txt", ".doc", ".docx", ".xlsx", ".csv"}


@dataclass
class ExtractedDoc:
    filename: str
    ext: str
    text: str


def _clean_text(s: str) -> str:
    return normalize_text(s)


def _extract_text_txt(data: bytes) -> str:
    # пробуем UTF-8, потом cp1251 (часто в РФ документах)
    for enc in ("utf-8", "utf-8-sig", "cp1251"):
        try:
            return data.decode(enc, errors="strict")
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")


def _extract_text_xlsx(data: bytes) -> str:
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")
    f = io.BytesIO(data)
    wb = openpyxl.load_workbook(f, data_only=True)
    out: List[str] = []
    for sheet in wb.worksheets:
        out.append(f"[SHEET] {sheet.title}")
        # аккуратно: ограничим “полотно”
        max_row = min(sheet.max_row or 0, 300)
        max_col = min(sheet.max_column or 0, 30)
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                v = sheet.cell(row=r, column=c).value
                if v is None:
                    row_vals.append("")
                else:
                    row_vals.append(str(v))
            # не тащим пустые строки
            if any(x.strip() for x in row_vals):
                out.append(" | ".join(row_vals))
        out.append("")  # пустая строка между листами
    return "\n".join(out).strip()


async def extract_docs_from_uploads(files: List[UploadFile]) -> List[ExtractedDoc]:
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")

    if len(files) > MAX_FILES:
        raise HTTPException(status_code=400, detail=f"Too many files. Max {MAX_FILES}")

    docs: List[ExtractedDoc] = []
    non_empty_count = 0

    for uf in files:
        filename = uf.filename or "file"
        ext = os.path.splitext(filename.lower())[1]

        if ext not in ALLOWED_EXT:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {ext}. Allowed: {sorted(ALLOWED_EXT)}"
            )

        data = await uf.read()
        if len(data) > MAX_FILE_BYTES:
            raise HTTPException(status_code=400, detail=f"File too large: {filename}")

        if ext in (".pdf", ".docx", ".doc"):
            text, _ = extract_text_from_document(filename, data)
        elif ext in (".txt", ".csv"):
            text = _extract_text_txt(data)
        elif ext == ".xlsx":
            text = _extract_text_xlsx(data)
        else:
            # на всякий
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")

        text = _clean_text(text)
        if text:
            non_empty_count += 1

        docs.append(ExtractedDoc(filename=filename, ext=ext, text=text))

    if not docs or non_empty_count == 0:
        raise HTTPException(status_code=400, detail="No text extracted from provided files")

    logger.info(
        "batch.docs.extracted",
        extra={
            "file_count": len(docs),
            "non_empty_count": non_empty_count,
            "files": [{"name": d.filename, "chars": len(d.text)} for d in docs],
        },
    )

    return docs


def build_structured_corpus(docs: List[ExtractedDoc]) -> str:
    # важное: явные границы файлов
    chunks: List[str] = []
    chunks.append("TENDER PACKAGE (MULTI-FILE) - STRUCTURED CORPUS\n")
    for i, d in enumerate(docs, start=1):
        chunks.append(f"\n===== FILE {i}/{len(docs)}: {d.filename} ({d.ext}) =====\n")
        chunks.append(d.text)
    corpus = normalize_text("\n".join(chunks))

    corpus_lower = corpus.lower()
    contract_files = [d.filename for d in docs if re.search(r"(договор|контракт|проект)", d.filename.lower())]
    tech_files = [d.filename for d in docs if re.search(r"(тех|техничес|тз|техзад|специф|requirements)", d.filename.lower())]

    if contract_files:
        if not any(k in corpus_lower for k in ("оплата", "неустойк", "штраф", "пеня")):
            logger.warning("contract_text_missing", extra={"files": contract_files})

    if tech_files:
        if not any(k in corpus_lower for k in ("срок", "партия", "отгруз", "поставка")):
            logger.warning("tech_text_missing", extra={"files": tech_files})

    logger.info(
        "batch.corpus.built",
        extra={"file_count": len(docs), "corpus_chars": len(corpus)},
    )

    return corpus
